import csv
from io import BytesIO, StringIO
from pathlib import PurePosixPath
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

SUPPORTED_FILE_TYPES = ["csv", "tsv", "txt", "xlsx"]
TEXT_FILE_TYPES = {"csv", "tsv", "txt"}
EXCEL_FILE_TYPES = {"xlsx"}
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_ROWS = 1_000_000
MAX_COLUMNS = 500
MAX_CELLS = 20_000_000
CSV_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")
MAX_XLSX_ARCHIVE_ENTRIES = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 256 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200.0
REQUIRED_XLSX_MEMBERS = {"[Content_Types].xml", "xl/workbook.xml"}


def _detect_delimiter(sample_text):
    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=",\t;|")
        return dialect.delimiter
    except csv.Error:
        return ","


def _decode_text_content(content):
    for encoding in ["utf-8-sig", "utf-8"]:
        try:
            return content.decode(encoding), None
        except UnicodeDecodeError:
            continue

    return (
        content.decode("latin-1"),
        "The file was decoded with a latin-1 fallback, so check special characters if text looks unusual.",
    )


def _build_schema_preview(df):
    rows = []
    row_count = max(len(df), 1)
    for column in df.columns:
        series = df[column]
        non_null = int(series.notna().sum())
        rows.append(
            {
                "column": str(column),
                "dtype": str(series.dtype),
                "non_null": non_null,
                "missing_pct": round(series.isna().mean() * 100, 1),
                "coverage_pct": round((non_null / row_count) * 100, 1),
                "unique_values": int(series.nunique(dropna=True)),
            }
        )
    return pd.DataFrame(rows)


def _normalize_and_validate_headers(headers):
    normalized = []
    for index, header in enumerate(headers):
        if pd.isna(header) or not str(header).strip():
            normalized.append(f"Unnamed: {index}")
        else:
            normalized.append(str(header).strip())

    duplicate_names = pd.Index(normalized)[pd.Index(normalized).duplicated()].unique().tolist()
    if duplicate_names:
        raise ValueError(
            "Column names must be unique after trimming whitespace. Duplicate columns: "
            + ", ".join(duplicate_names[:5])
        )
    return normalized


def _validate_table_shape(df):
    rows, columns = df.shape
    if rows > MAX_ROWS:
        raise ValueError(f"The dataset has {rows:,} rows; the limit is {MAX_ROWS:,}.")
    if columns > MAX_COLUMNS:
        raise ValueError(f"The dataset has {columns:,} columns; the limit is {MAX_COLUMNS:,}.")
    if rows * columns > MAX_CELLS:
        raise ValueError(
            f"The dataset has {rows * columns:,} cells; the limit is {MAX_CELLS:,}."
        )


def _validate_shape_values(rows, columns):
    if rows > MAX_ROWS:
        raise ValueError(f"The dataset has {rows:,} rows; the limit is {MAX_ROWS:,}.")
    if columns > MAX_COLUMNS:
        raise ValueError(f"The dataset has {columns:,} columns; the limit is {MAX_COLUMNS:,}.")
    if rows * columns > MAX_CELLS:
        raise ValueError(
            f"The dataset has {rows * columns:,} cells; the limit is {MAX_CELLS:,}."
        )


def _preflight_xlsx_archive(content):
    """Reject suspicious XLSX archives before an XML/workbook parser sees them."""
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise ValueError(
                    f"The Excel archive has {len(members):,} entries; the safety limit is "
                    f"{MAX_XLSX_ARCHIVE_ENTRIES:,}."
                )

            names = set()
            total_uncompressed = 0
            for member in members:
                path = PurePosixPath(member.filename)
                if path.is_absolute() or ".." in path.parts:
                    raise ValueError("The Excel archive contains an unsafe member path.")
                if member.filename in names:
                    raise ValueError("The Excel archive contains duplicate member names.")
                names.add(member.filename)

                if member.flag_bits & 0x1:
                    raise ValueError("Encrypted Excel workbooks are not supported.")
                if member.is_dir():
                    continue

                if member.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise ValueError(
                        "An Excel archive entry exceeds the maximum uncompressed member size "
                        f"of {MAX_XLSX_MEMBER_BYTES / (1024 * 1024):.0f} MB."
                    )
                total_uncompressed += member.file_size
                if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError(
                        "The Excel archive's uncompressed size exceeds the safety limit of "
                        f"{MAX_XLSX_UNCOMPRESSED_BYTES / (1024 * 1024):.0f} MB."
                    )

                if member.file_size:
                    if not member.compress_size:
                        raise ValueError(
                            "The Excel archive contains an invalid compressed entry."
                        )
                    ratio = member.file_size / member.compress_size
                    if ratio > MAX_XLSX_COMPRESSION_RATIO:
                        raise ValueError(
                            "The Excel archive contains a suspicious compression ratio "
                            f"({ratio:.1f}:1; limit {MAX_XLSX_COMPRESSION_RATIO:.1f}:1)."
                        )

            if not REQUIRED_XLSX_MEMBERS.issubset(names):
                raise ValueError("The upload is not a valid XLSX workbook archive.")
    except BadZipFile as exc:
        raise ValueError("The upload is not a valid XLSX workbook archive.") from exc


def _read_xlsx_streaming(content):
    """Read the first worksheet with resource limits enforced during iteration."""
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        OSError,
        ParseError,
        ValueError,
    ) as exc:
        raise ValueError("The upload is not a readable XLSX workbook.") from exc

    try:
        sheet_names = workbook.sheetnames
        if not sheet_names:
            raise ValueError("The Excel workbook has no worksheets.")

        sheet_name = sheet_names[0]
        worksheet = workbook[sheet_name]
        declared_rows = int(worksheet.max_row or 0)
        declared_columns = int(worksheet.max_column or 0)
        _validate_shape_values(max(declared_rows - 1, 0), declared_columns)
        # Worksheet dimensions are untrusted XML metadata and may understate the
        # real used range. Reset them so streaming discovers every stored cell;
        # the limits below are enforced as those cells are encountered.
        worksheet.reset_dimensions()

        row_iterator = worksheet.iter_rows(values_only=True)
        raw_header = list(next(row_iterator, ()))
        column_values = [[] for _ in raw_header]
        rows_seen = 0
        last_nonempty_row = 0

        for raw_row in row_iterator:
            row = list(raw_row)
            rows_seen += 1
            width = max(len(raw_header), len(row))
            _validate_shape_values(rows_seen, width)

            if width > len(raw_header):
                new_columns = width - len(raw_header)
                raw_header.extend([None] * new_columns)
                column_values.extend([[None] * (rows_seen - 1) for _ in range(new_columns)])
            if len(row) < width:
                row.extend([None] * (width - len(row)))

            for index, value in enumerate(row):
                column_values[index].append(value)
            if any(value is not None for value in row):
                last_nonempty_row = rows_seen

        if not raw_header or (
            not any(value is not None for value in raw_header) and not last_nonempty_row
        ):
            return pd.DataFrame(), sheet_name, sheet_names

        headers = _normalize_and_validate_headers(raw_header)
        trimmed_columns = {
            header: values[:last_nonempty_row]
            for header, values in zip(headers, column_values, strict=True)
        }
        dataframe = pd.DataFrame(trimmed_columns)
        _validate_table_shape(dataframe)
        return dataframe, sheet_name, sheet_names
    finally:
        workbook.close()


def read_uploaded_table_details(uploaded_file):
    file_name = uploaded_file.name
    suffix = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""

    content = uploaded_file.getvalue()
    if not content:
        raise ValueError("The uploaded file is empty.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"The uploaded file is {len(content) / (1024 * 1024):.1f} MB; the limit is "
            f"{MAX_UPLOAD_BYTES // (1024 * 1024)} MB."
        )

    if suffix not in SUPPORTED_FILE_TYPES:
        raise ValueError(
            f"Unsupported file type '{suffix}'. Please upload one of: "
            + ", ".join(SUPPORTED_FILE_TYPES)
        )

    warnings = []
    delimiter = None
    sheet_name = None

    if suffix in TEXT_FILE_TYPES:
        decoded, decode_warning = _decode_text_content(content)
        if decode_warning:
            warnings.append(decode_warning)
        delimiter = _detect_delimiter(decoded[:5000])
        header_row = next(csv.reader(StringIO(decoded), delimiter=delimiter), [])
        _normalize_and_validate_headers(header_row)
        df = pd.read_csv(StringIO(decoded), sep=delimiter)
        source_summary = f"Loaded {suffix.upper()} file with '{delimiter}' as the detected delimiter."
    else:
        _preflight_xlsx_archive(content)
        df, sheet_name, sheet_names = _read_xlsx_streaming(content)
        if len(sheet_names) > 1:
            warnings.append(
                f"Workbook has {len(sheet_names)} sheets. The app used the first sheet: '{sheet_name}'."
            )
        source_summary = f"Loaded Excel workbook from sheet '{sheet_name}'."

    df.columns = _normalize_and_validate_headers(df.columns)
    _validate_table_shape(df)

    if df.empty:
        warnings.append("The uploaded table has headers but no data rows.")

    return {
        "dataframe": df,
        "file_type": suffix,
        "source_summary": source_summary,
        "delimiter": delimiter,
        "sheet_name": sheet_name,
        "warnings": warnings,
        "schema_preview": _build_schema_preview(df),
    }


def read_uploaded_table(uploaded_file):
    return read_uploaded_table_details(uploaded_file)["dataframe"]


def dataframe_to_csv_bytes(df):
    safe = df.copy()

    def escape_formula(value):
        if isinstance(value, str) and value.startswith(CSV_FORMULA_PREFIXES):
            return "'" + value
        return value

    def safe_headers(columns):
        exported = []
        used = set()
        for index, column in enumerate(columns):
            header = "" if column is None else str(column)
            header = escape_formula(header)
            if not header:
                header = f"Unnamed: {index}"

            candidate = header
            suffix = 2
            while candidate in used:
                candidate = f"{header}__{suffix}"
                suffix += 1
            exported.append(candidate)
            used.add(candidate)
        return exported

    safe.columns = safe_headers(safe.columns)

    for column in safe.columns:
        if (
            pd.api.types.is_object_dtype(safe[column])
            or pd.api.types.is_string_dtype(safe[column])
            or isinstance(safe[column].dtype, pd.CategoricalDtype)
        ):
            safe[column] = safe[column].map(escape_formula)

    # CRLF makes Python's CSV writer quote fields containing either CR or LF.
    # Without it, a carriage return can split a record and expose a formula on
    # the next physical line despite the leading apostrophe.
    return safe.to_csv(index=False, lineterminator="\r\n").encode("utf-8")
